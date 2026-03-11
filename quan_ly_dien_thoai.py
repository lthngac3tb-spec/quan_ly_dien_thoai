import streamlit as st
import pandas as pd
from datetime import datetime
import pytz
import os
import io

# --- 1. CẤU HÌNH HỆ THỐNG ---
FILE_LOP = "danh_sach_lop.xlsx"
mui_gio_vn = pytz.timezone('Asia/Ho_Chi_Minh')

st.set_page_config(page_title="Quản lý Điện thoại", page_icon="📱", layout="centered")
st.title("📱 HỆ THỐNG QUẢN LÝ ĐIỆN THOẠI")

# --- 2. HÀM ĐỌC DỮ LIỆU ---
def load_data():
    if os.path.exists(FILE_LOP):
        df = pd.read_excel(FILE_LOP)
        df['STT'] = df['STT'].astype(str).apply(lambda x: x.split('.')[0].strip().zfill(2))
        # Chuẩn hóa tất cả các cột về string, bỏ khoảng trắng thừa
        for col in ['HoTen', 'TrangThai', 'GioCat', 'GioTra']:
            if col in df.columns:
                df[col] = df[col].astype(str).replace('nan', '').str.strip()
        return df
    else:
        st.error(f"Không tìm thấy file {FILE_LOP}!")
        return None

df = load_data()

if df is not None:
    tab_thu, tab_tra, tab_bc = st.tabs(["📥 THU MÁY", "📤 TRẢ MÁY", "📊 BÁO CÁO"])

    # --- TAB 1: THU MÁY ---
    with tab_thu:
        st.subheader("📥 Trạm Thu Máy (Đầu giờ)")
        with st.form(key='form_thu_may', clear_on_submit=True):
            ma_thu = st.text_input("Nhập STT bạn nộp máy rồi Enter:")
            submit_thu = st.form_submit_button(label='Xác nhận Thu')
        
        if submit_thu and ma_thu:
            stt_nhan = ma_thu.strip().zfill(2)
            if stt_nhan in df['STT'].values:
                idx = df.index[df['STT'] == stt_nhan][0]
                trang_thai = df.at[idx, 'TrangThai']
                ten_hs = df.at[idx, 'HoTen']

                if trang_thai == "✅ Đã cất":
                    st.warning(f"⚠️ {ten_hs} đã nộp máy rồi.")
                else:
                    df.at[idx, 'TrangThai'] = "✅ Đã cất"
                    df.at[idx, 'GioCat'] = datetime.now(mui_gio_vn).strftime("%H:%M %d/%m")
                    df.to_excel(FILE_LOP, index=False)
                    st.success(f"✅ Đã thu máy của: {ten_hs}")
            else:
                st.warning(f"❌ Không thấy STT: {stt_nhan}")

    # --- TAB 2: TRẢ MÁY (SỬA LỖI CHẶN TRẢ) ---
    with tab_tra:
        st.subheader("📤 Trạm Trả Máy (Tan học)")
        with st.form(key='form_tra_may', clear_on_submit=True):
            ma_tra = st.text_input("Nhập STT bạn lấy máy rồi Enter:")
            submit_tra = st.form_submit_button(label='Xác nhận Trả')
        
        if submit_tra and ma_tra:
            stt_nhan = ma_tra.strip().zfill(2)
            if stt_nhan in df['STT'].values:
                idx = df.index[df['STT'] == stt_nhan][0]
                trang_thai = df.at[idx, 'TrangThai']
                ten_hs = df.at[idx, 'HoTen']

                # ĐIỀU KIỆN CHẶN CHẶT CHẼ:
                # Chỉ cho phép trả nếu trạng thái CHÍNH XÁC là "✅ Đã cất"
                if trang_thai == "✅ Đã cất":
                    df.at[idx, 'TrangThai'] = "🏠 Đã trả"
                    df.at[idx, 'GioTra'] = datetime.now(mui_gio_vn).strftime("%H:%M %d/%m")
                    df.to_excel(FILE_LOP, index=False)
                    st.info(f"🏠 Đã trả máy cho: {ten_hs}")
                elif trang_thai == "🏠 Đã trả":
                    st.warning(f"⚠️ {ten_hs} đã nhận máy trước đó rồi.")
                else:
                    # Trường hợp "Chưa cất", "nan", hoặc rỗng
                    st.error(f"🚫 KHÔNG THỂ TRẢ: Bạn {ten_hs} chưa nộp máy vào tủ!")
            else:
                st.warning(f"❌ Không thấy STT: {stt_nhan}")

    # --- TAB 3: BÁO CÁO ---
    with tab_bc:
        col1, col2 = st.columns(2)
        da_cat = len(df[df['TrangThai'] == "✅ Đã cất"])
        col1.metric("Tổng sĩ số", f"{len(df)} HS")
        col2.metric("Máy đang giữ", f"{da_cat} máy")
        st.divider()
        st.dataframe(df, use_container_width=True)

        # XUẤT FILE EXCEL
        try:
            from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
            from openpyxl.utils import get_column_letter
            from openpyxl import Workbook

            if st.button("📝 Chuẩn bị file báo cáo"):
                ngay_hien_tai = datetime.now(mui_gio_vn).strftime("%d_%m_%Y")
                buffer = io.BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "BaoCao"
                headers = list(df.columns)
                ws.append(headers)
                for r in df.values.tolist():
                    row_data = [str(x) if str(x) != 'nan' else "" for x in r]
                    ws.append(row_data)
                
                # Căn chỉnh và kẻ bảng
                blue_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.fill = blue_fill
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.border = thin_border

                column_widths = [10, 25, 15, 15, 15]
                for i, width in enumerate(column_widths):
                    ws.column_dimensions[get_column_letter(i + 1)].width = width

                wb.save(buffer)
                # --- 3. QUẢN LÝ DỮ LIỆU & XUẤT FILE ĐẸP ---
        st.subheader("💾 Công cụ quản lý")
        
        # Tạo tên file có kèm ngày hiện tại
        ngay_hien_tai = datetime.now().strftime("%d-%m-%Y")
        ten_file_xuat = f"Danh_sach_khach_{ngay_hien_tai}.xlsx"
        
        # --- LOGIC TẠO FILE EXCEL ĐỊNH DẠNG ĐẸP ---
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
            # Xuất dữ liệu ra sheet
            df.to_excel(writer, index=False, sheet_name='DanhSachKhach')
            
            workbook  = writer.book
            worksheet = writer.sheets['DanhSachKhach']

            # 1. Định dạng tiêu đề (Header): Màu nền xanh, chữ trắng, in đậm, căn giữa, kẻ bảng
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'vcenter',
                'align': 'center',
                'fg_color': '#D7E4BC',
                'border': 1
            })

            # 2. Định dạng nội dung (Data): Kẻ bảng, căn giữa vcenter
            cell_format = workbook.add_format({
                'border': 1,
                'valign': 'vcenter'
            })

            # Ghi đè tiêu đề với định dạng mới
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
            
            # Ghi định dạng cho toàn bộ các ô dữ liệu
            for row_num in range(1, len(df) + 1):
                for col_num in range(len(df.columns)):
                    # Lấy dữ liệu hiện tại để ghi lại với format
                    val = df.iloc[row_num-1, col_num] if row_num <= len(df) else ""
                    worksheet.write(row_num, col_num, val, cell_format)

            # 3. Tự động chỉnh độ rộng cột cho vừa nội dung
            for i, col in enumerate(df.columns):
                column_len = max(df[col].astype(str).str.len().max(), len(col)) + 2
                worksheet.set_column(i, i, column_len)

        # Nút tải file
        st.download_button(
            label="📥 Tải File Excel Báo Cáo (Đã kẻ bảng)",
            data=buffer.getvalue(),
            file_name=ten_file_xuat,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        except Exception as e:
            st.error(f"Lỗi: {e}")

        if st.button("🔄 Reset dữ liệu ngày mới"):
            df['TrangThai'] = "Chưa nộp"
            df['GioCat'] = ""
            df['GioTra'] = ""
            df.to_excel(FILE_LOP, index=False)
            st.rerun()

